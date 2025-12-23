import os
import sys
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
import sqlite3
from dotenv import load_dotenv
import schedule
import time

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

logger = logging.getLogger(__name__)



class ProcesadorVentas:
    def __init__(self):
        load_dotenv()

        ruta_base_datos = os.getenv('DB_PATH', 'anepsa_ventas.db')
        self.ruta_base_datos = Path(__file__).parent / ruta_base_datos

        self.ruta_archivo_csv_entrada = Path(__file__).parent / "Archivo ventas_simuladas.csv"
        self.ruta_archivo_excel_salida = Path(__file__).parent / "resumen_ventas_mensual.xlsx"
        self.nombre_tabla = "ventas_raw"

    def obtener_conexion(self):
        try:
            conexion = sqlite3.connect(self.ruta_base_datos)
            logger.info(f"Conexión a base de datos SQLite establecida: {self.ruta_base_datos}")
            return conexion
        except sqlite3.Error as error:
            logger.error(f"Error al conectar a la base de datos: {error}")
            raise

    def crear_tabla_si_no_existe(self, conexion):
        logger.info(f"Verificando existencia de tabla '{self.nombre_tabla}'")

        consulta_crear_tabla = f"""
        CREATE TABLE IF NOT EXISTS {self.nombre_tabla} (
            fecha TEXT NOT NULL,
            region TEXT NOT NULL,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio_unitario REAL NOT NULL
        )
        """

        try:
            cursor = conexion.cursor()
            cursor.execute(consulta_crear_tabla)
            conexion.commit()
            logger.info(f"Tabla '{self.nombre_tabla}' verificada/creada exitosamente")
            cursor.close()
        except sqlite3.Error as error:
            logger.error(f"Error al crear tabla: {error}")
            conexion.rollback()
            raise

    def limpiar_tabla(self, conexion):
        logger.info(f"Limpiando tabla '{self.nombre_tabla}'")

        try:
            cursor = conexion.cursor()
            cursor.execute(f"DELETE FROM {self.nombre_tabla}")
            conexion.commit()
            logger.info("Tabla limpiada exitosamente")
            cursor.close()
        except sqlite3.Error as error:
            logger.error(f"Error al limpiar tabla: {error}")
            conexion.rollback()
            raise

    def cargar_datos_desde_csv(self):
        logger.info(f"Cargando datos desde '{self.ruta_archivo_csv_entrada}'")

        if not self.ruta_archivo_csv_entrada.exists():
            logger.error(f"Archivo CSV no encontrado: {self.ruta_archivo_csv_entrada}")
            raise FileNotFoundError(f"Archivo no encontrado: {self.ruta_archivo_csv_entrada}")

        try:
            dataframe_ventas = pd.read_csv(self.ruta_archivo_csv_entrada)
            cantidad_registros_inicial = len(dataframe_ventas)
            logger.info(f"Datos cargados: {cantidad_registros_inicial} registros")

            return dataframe_ventas
        except Exception as error:
            logger.error(f"Error al cargar CSV: {error}")
            raise

    def insertar_datos_en_tabla(self, conexion, dataframe_ventas):
        logger.info("Insertando datos en la base de datos")

        try:
            cursor = conexion.cursor()

            consulta_insercion = f"""
            INSERT INTO {self.nombre_tabla}
            (fecha, region, producto, cantidad, precio_unitario)
            VALUES (?, ?, ?, ?, ?)
            """

            registros_insertados = 0

            for indice, fila in dataframe_ventas.iterrows():
                cursor.execute(
                    consulta_insercion,
                    (
                        fila['fecha'],
                        fila['region'],
                        fila['producto'],
                        fila['cantidad'],
                        fila['precio_unitario']
                    )
                )
                registros_insertados += 1

            conexion.commit()
            logger.info(f"Insertados {registros_insertados} registros en la base de datos")
            cursor.close()

        except sqlite3.Error as error:
            logger.error(f"Error al insertar datos: {error}")
            conexion.rollback()
            raise

    def eliminar_duplicados(self, conexion):
        logger.info("Eliminando registros duplicados")

        consulta_eliminar_duplicados = f"""
        DELETE FROM {self.nombre_tabla}
        WHERE rowid NOT IN (
            SELECT MIN(rowid)
            FROM {self.nombre_tabla}
            GROUP BY fecha, region, producto, cantidad, precio_unitario
        )
        """

        try:
            cursor = conexion.cursor()
            cursor.execute(consulta_eliminar_duplicados)
            cantidad_duplicados_eliminados = cursor.rowcount
            conexion.commit()
            logger.info(f"Eliminados {cantidad_duplicados_eliminados} registros duplicados")
            cursor.close()

            return cantidad_duplicados_eliminados

        except sqlite3.Error as error:
            logger.error(f"Error al eliminar duplicados: {error}")
            conexion.rollback()
            raise

    def eliminar_registros_con_nulos(self, conexion):
        logger.info("Eliminando registros con valores nulos")

        consulta_eliminar_nulos = f"""
        DELETE FROM {self.nombre_tabla}
        WHERE fecha IS NULL
           OR region IS NULL
           OR producto IS NULL
           OR cantidad IS NULL
           OR precio_unitario IS NULL
        """

        try:
            cursor = conexion.cursor()
            cursor.execute(consulta_eliminar_nulos)
            cantidad_nulos_eliminados = cursor.rowcount
            conexion.commit()
            logger.info(f"Eliminados {cantidad_nulos_eliminados} registros con valores nulos")
            cursor.close()

            return cantidad_nulos_eliminados

        except sqlite3.Error as error:
            logger.error(f"Error al eliminar registros nulos: {error}")
            conexion.rollback()
            raise

    def calcular_resumen_ventas(self, conexion):
        logger.info("Calculando resumen de ventas por región y mes")

        consulta_resumen = f"""
        SELECT
            region,
            CAST(strftime('%Y', fecha) AS INTEGER) as año,
            CAST(strftime('%m', fecha) AS INTEGER) as mes,
            SUM(cantidad * precio_unitario) as total_ventas
        FROM {self.nombre_tabla}
        GROUP BY region, año, mes
        ORDER BY region, año, mes
        """

        try:
            dataframe_resumen = pd.read_sql_query(consulta_resumen, conexion)

            dataframe_resumen['total_ventas'] = dataframe_resumen['total_ventas'].round(2)

            cantidad_registros_resumen = len(dataframe_resumen)
            logger.info(f"Resumen calculado: {cantidad_registros_resumen} registros")

            return dataframe_resumen

        except Exception as error:
            logger.error(f"Error al calcular resumen: {error}")
            raise

    def exportar_resumen_a_excel(self, dataframe_resumen):
        logger.info(f"Exportando resumen a '{self.ruta_archivo_excel_salida}'")

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font

            dataframe_resumen.to_excel(
                self.ruta_archivo_excel_salida,
                index=False,
                engine='openpyxl'
            )

            workbook = load_workbook(self.ruta_archivo_excel_salida)
            worksheet = workbook.active

            color_encabezado = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            fuente_encabezado = Font(bold=True, color="FFFFFF")

            for celda in worksheet[1]:
                celda.fill = color_encabezado
                celda.font = fuente_encabezado

            workbook.save(self.ruta_archivo_excel_salida)
            logger.info("Resumen exportado exitosamente")

        except Exception as error:
            logger.error(f"Error al exportar resumen: {error}")
            raise

    def procesar_datos_completo(self):
        logger.info("=" * 60)
        logger.info("INICIANDO PROCESAMIENTO DE DATOS DE VENTAS")
        logger.info("=" * 60)

        hora_inicio = datetime.now()

        try:
            conexion = self.obtener_conexion()

            self.crear_tabla_si_no_existe(conexion)

            self.limpiar_tabla(conexion)

            dataframe_ventas = self.cargar_datos_desde_csv()

            self.insertar_datos_en_tabla(conexion, dataframe_ventas)

            cantidad_duplicados = self.eliminar_duplicados(conexion)

            cantidad_nulos = self.eliminar_registros_con_nulos(conexion)

            dataframe_resumen = self.calcular_resumen_ventas(conexion)

            self.exportar_resumen_a_excel(dataframe_resumen)

            conexion.close()
            logger.info("Conexión a base de datos cerrada")

            hora_fin = datetime.now()
            duracion_total = (hora_fin - hora_inicio).total_seconds()

            logger.info("=" * 60)
            logger.info("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
            logger.info(f"Duplicados eliminados: {cantidad_duplicados}")
            logger.info(f"Registros con nulos eliminados: {cantidad_nulos}")
            logger.info(f"Registros en resumen final: {len(dataframe_resumen)}")
            logger.info(f"Tiempo total de procesamiento: {duracion_total:.2f} segundos")
            logger.info("=" * 60)

        except Exception as error:
            logger.error("=" * 60)
            logger.error("ERROR EN EL PROCESAMIENTO")
            logger.error(f"Tipo de error: {type(error).__name__}")
            logger.error(f"Detalle: {error}")
            logger.error("=" * 60)
            raise


def ejecutar_tarea_programada():
    logger.info("Ejecutando tarea programada a las 08:00 AM")
    procesador = ProcesadorVentas()
    procesador.procesar_datos_completo()


def configurar_ejecucion_automatica():
    logger.info("Configurando ejecución automática diaria a las 08:00 AM")
    schedule.every().day.at("08:00").do(ejecutar_tarea_programada)

    logger.info("Ejecutando primera vez inmediatamente")
    ejecutar_tarea_programada()

    logger.info("Esperando próxima ejecución programada...")

    while True:
        schedule.run_pending()
        time.sleep(60)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--automatico":
        configurar_ejecucion_automatica()
    else:
        procesador = ProcesadorVentas()
        procesador.procesar_datos_completo()
